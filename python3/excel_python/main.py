from data_json import data
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


try:
    #datos en json o diccionario
    myData = data

    #ajustes del archivo .excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas" #nombre de la hoja de excel

    headings = ['Name'] + list(data['Yeisson'].keys()) #cabeceras del excel
    ws.append(headings) #a;adir cabeceras a mi archivo python excel


    #recorro mi archivo json por persona con un for
    for name in data:
        notas_por_persona = list(data[name].values()) # datos de la fila, persona x
        ws.append([name] + notas_por_persona) # agrego a la persona con sus notas, fila por fila


    #promedio de notas fila ultima 
    for col in range(2, len(data['Esteban']) + 2):
        char = get_column_letter(col)
        ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"
    
    #estilos a la hoja
    for col in range(1, 6):
        ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

    #guardo el archivo
    wb.save("Notas_por_persona3.xlsx")
except:
    print('simple try catch de errores: si se muestra esto es porque algo se ha codificado mal')